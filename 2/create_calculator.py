#!/usr/bin/env python3
"""
Калькулятор расходов на AI-проект (Fashion Try-On)
Создает Excel файл с 6 листами и формулами
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule

# Создаем книгу
wb = Workbook()

# Цвета
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
SECTION_FILL = PatternFill(start_color="D6DCE5", end_color="D6DCE5", fill_type="solid")
TOTAL_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
INPUT_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

def style_header(ws, row, cols):
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = BORDER

def style_cell(ws, row, col, fill=None):
    cell = ws.cell(row=row, column=col)
    cell.border = BORDER
    if fill:
        cell.fill = fill
    return cell

def set_column_widths(ws, widths):
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

# ==================== ЛИСТ 1: CONTROL ====================
ws_control = wb.active
ws_control.title = "CONTROL"

ws_control['A1'] = "ПАНЕЛЬ УПРАВЛЕНИЯ - ВВОДНЫЕ ДАННЫЕ"
ws_control.merge_cells('A1:D1')
ws_control['A1'].font = Font(bold=True, size=14)
ws_control['A1'].fill = HEADER_FILL
ws_control['A1'].font = Font(color="FFFFFF", bold=True, size=14)

headers = ["Параметр", "Значение", "Единица", "Комментарий"]
for col, header in enumerate(headers, 1):
    ws_control.cell(row=3, column=col, value=header)
style_header(ws_control, 3, 4)

# Данные CONTROL
control_data = [
    # row 4-17
    ("ПОЛЬЗОВАТЕЛИ", "", "", ""),
    ("Пользователи (месяц 1)", 5000, "чел", "MVP старт"),
    ("Рост пользователей", 15, "%", "Ежемесячный прирост"),
    ("Текущий месяц расчета", 1, "мес", "Меняй для прогноза (1-12)"),
    ("", "", "", ""),
    ("ИСПОЛЬЗОВАНИЕ НА ПОЛЬЗОВАТЕЛЯ", "", "", ""),
    ("Примерок на пользователя", 3, "шт", "Pixelcut Try-On"),
    ("Анимаций (видео) на пользователя", 1, "шт", "Video generation"),
    ("LLM запросов на пользователя", 5, "шт", "Чат/рекомендации"),
    ("", "", "", ""),
    ("РАЗМЕРЫ ФАЙЛОВ", "", "", ""),
    ("Средний размер фото", 3, "MB", "Входное изображение"),
    ("Средний размер видео", 15, "MB", "Сгенерированное видео"),
    ("", "", "", ""),
    ("МАСШТАБИРОВАНИЕ", "", "", ""),
    ("Пользователей на 1 сервер", 10000, "чел", "Порог для автоскейла"),
]

for i, (param, value, unit, comment) in enumerate(control_data, 4):
    ws_control.cell(row=i, column=1, value=param)
    cell_val = ws_control.cell(row=i, column=2, value=value)
    ws_control.cell(row=i, column=3, value=unit)
    ws_control.cell(row=i, column=4, value=comment)

    # Стилизация секций
    if param in ["ПОЛЬЗОВАТЕЛИ", "ИСПОЛЬЗОВАНИЕ НА ПОЛЬЗОВАТЕЛЯ", "РАЗМЕРЫ ФАЙЛОВ", "МАСШТАБИРОВАНИЕ"]:
        for col in range(1, 5):
            style_cell(ws_control, i, col, SECTION_FILL)
    elif value != "":
        style_cell(ws_control, i, 2, INPUT_FILL)

    for col in range(1, 5):
        ws_control.cell(row=i, column=col).border = BORDER

# Расчетные поля
ws_control.cell(row=21, column=1, value="РАСЧЕТНЫЕ ЗНАЧЕНИЯ")
ws_control.merge_cells('A21:D21')
style_cell(ws_control, 21, 1, HEADER_FILL)
ws_control['A21'].font = HEADER_FONT

calc_data = [
    ("Пользователи (текущий месяц)", "=IF(B7=1,B5,ROUND(B5*(1+B6/100)^(B7-1),0))", "чел", "С учетом роста"),
    ("Всего примерок", "=B22*B10", "шт", "В месяц"),
    ("Всего видео", "=B22*B11", "шт", "В месяц"),
    ("Всего LLM запросов", "=B22*B12", "шт", "В месяц"),
    ("Хранилище фото (GB)", "=B23*B15/1024", "GB", "В месяц"),
    ("Хранилище видео (GB)", "=B24*B16/1024", "GB", "В месяц"),
    ("Требуется серверов", "=CEILING(B22/B19,1)", "шт", "Автоскейл"),
]

for i, (param, formula, unit, comment) in enumerate(calc_data, 22):
    ws_control.cell(row=i, column=1, value=param)
    ws_control.cell(row=i, column=2, value=formula)
    ws_control.cell(row=i, column=3, value=unit)
    ws_control.cell(row=i, column=4, value=comment)
    style_cell(ws_control, i, 2, TOTAL_FILL)
    for col in range(1, 5):
        ws_control.cell(row=i, column=col).border = BORDER

set_column_widths(ws_control, [35, 20, 10, 35])

# ==================== ЛИСТ 2: ИИ И ГЕНЕРАЦИЯ ====================
ws_ai = wb.create_sheet("AI_Generation")

ws_ai['A1'] = "ИИ И ГЕНЕРАЦИЯ - ЗАТРАТЫ"
ws_ai.merge_cells('A1:E1')
ws_ai['A1'].font = Font(bold=True, size=14, color="FFFFFF")
ws_ai['A1'].fill = HEADER_FILL

headers = ["Сервис / Операция", "Цена за 1 опер. ($)", "Количество", "Итого ($)", "Комментарий"]
for col, header in enumerate(headers, 1):
    ws_ai.cell(row=3, column=col, value=header)
style_header(ws_ai, 3, 5)

ai_data = [
    ("ВИРТУАЛЬНЫЕ ПРИМЕРКИ", "", "", "", ""),
    ("Pixelcut Try-On API", 0.10, "=CONTROL!B23", "=B5*C5", "10 кредитов = $0.10/изображение"),
    ("", "", "", "", ""),
    ("ВИДЕО ГЕНЕРАЦИЯ", "", "", "", ""),
    ("Video Animation (MiniMax/Runway)", 0.50, "=CONTROL!B24", "=B9*C9", "~$0.50 за 5 сек видео"),
    ("", "", "", "", ""),
    ("LLM / ТЕКСТ", "", "", "", ""),
    ("OpenAI GPT-4o (1K токенов)", 0.005, "=CONTROL!B25*2", "=B13*C13", "input+output ~2K токенов/запрос"),
    ("OpenAI GPT-4o-mini (backup)", 0.0003, "=CONTROL!B25*0.5", "=B14*C14", "Легкие запросы"),
    ("Embeddings (поиск)", 0.0001, "=CONTROL!B25", "=B15*C15", "Рекомендации"),
    ("", "", "", "", ""),
    ("ДОПОЛНИТЕЛЬНЫЕ AI", "", "", "", ""),
    ("Background Removal", 0.02, "=CONTROL!B23*0.5", "=B19*C19", "50% примерок"),
    ("Image Upscale", 0.01, "=CONTROL!B23*0.3", "=B20*C20", "30% примерок"),
]

for i, (service, price, qty, total, comment) in enumerate(ai_data, 4):
    ws_ai.cell(row=i, column=1, value=service)
    ws_ai.cell(row=i, column=2, value=price)
    ws_ai.cell(row=i, column=3, value=qty)
    ws_ai.cell(row=i, column=4, value=total)
    ws_ai.cell(row=i, column=5, value=comment)

    if service in ["ВИРТУАЛЬНЫЕ ПРИМЕРКИ", "ВИДЕО ГЕНЕРАЦИЯ", "LLM / ТЕКСТ", "ДОПОЛНИТЕЛЬНЫЕ AI"]:
        for col in range(1, 6):
            style_cell(ws_ai, i, col, SECTION_FILL)

    for col in range(1, 6):
        ws_ai.cell(row=i, column=col).border = BORDER

# Итого AI
ws_ai.cell(row=22, column=1, value="ИТОГО AI И ГЕНЕРАЦИЯ")
ws_ai.cell(row=22, column=4, value="=SUM(D5:D21)")
style_cell(ws_ai, 22, 1, TOTAL_FILL)
style_cell(ws_ai, 22, 4, TOTAL_FILL)
ws_ai['A22'].font = Font(bold=True)
ws_ai['D22'].font = Font(bold=True)
for col in range(1, 6):
    ws_ai.cell(row=22, column=col).border = BORDER

set_column_widths(ws_ai, [35, 20, 15, 15, 40])

# ==================== ЛИСТ 3: ИНФРАСТРУКТУРА ====================
ws_infra = wb.create_sheet("Infrastructure")

ws_infra['A1'] = "ИНФРАСТРУКТУРА - СЕРВЕРЫ И СЕРВИСЫ"
ws_infra.merge_cells('A1:E1')
ws_infra['A1'].font = Font(bold=True, size=14, color="FFFFFF")
ws_infra['A1'].fill = HEADER_FILL

headers = ["Ресурс", "Цена/мес ($)", "Количество", "Итого ($)", "Провайдер / Комментарий"]
for col, header in enumerate(headers, 1):
    ws_infra.cell(row=3, column=col, value=header)
style_header(ws_infra, 3, 5)

infra_data = [
    ("СЕРВЕРЫ", "", "", "", ""),
    ("Backend VPS (4 CPU / 8 GB)", 45, "=CONTROL!B28", "=B5*C5", "DigitalOcean / Vultr"),
    ("Worker Server (AI Queue)", 60, "=CEILING(CONTROL!B28/2,1)", "=B6*C6", "Для обработки задач"),
    ("", "", "", "", ""),
    ("БАЗЫ ДАННЫХ", "", "", "", ""),
    ("PostgreSQL Managed", 25, 1, "=B10*C10", "DigitalOcean / Supabase"),
    ("Redis (кэш/очереди)", 15, 1, "=B11*C11", "Upstash / Redis Cloud"),
    ("", "", "", "", ""),
    ("ХРАНИЛИЩЕ", "", "", "", ""),
    ("Object Storage (за GB)", 0.02, "=CONTROL!B26+CONTROL!B27", "=B15*C15", "S3 / Cloudflare R2"),
    ("CDN Bandwidth (за GB)", 0.01, "=(CONTROL!B26+CONTROL!B27)*3", "=B16*C16", "x3 от хранилища"),
    ("", "", "", "", ""),
    ("ДОПОЛНИТЕЛЬНО", "", "", "", ""),
    ("SSL / Domain", 0, 1, "=B20*C20", "Let's Encrypt бесплатно"),
    ("Monitoring (Sentry)", 26, 1, "=B21*C21", "Team plan"),
    ("CI/CD (GitHub Actions)", 0, 1, "=B22*C22", "Free tier достаточно"),
    ("Email Service (Resend)", 20, 1, "=B23*C23", "10K emails/month"),
]

for i, (resource, price, qty, total, comment) in enumerate(infra_data, 4):
    ws_infra.cell(row=i, column=1, value=resource)
    ws_infra.cell(row=i, column=2, value=price)
    ws_infra.cell(row=i, column=3, value=qty)
    ws_infra.cell(row=i, column=4, value=total)
    ws_infra.cell(row=i, column=5, value=comment)

    if resource in ["СЕРВЕРЫ", "БАЗЫ ДАННЫХ", "ХРАНИЛИЩЕ", "ДОПОЛНИТЕЛЬНО"]:
        for col in range(1, 6):
            style_cell(ws_infra, i, col, SECTION_FILL)

    for col in range(1, 6):
        ws_infra.cell(row=i, column=col).border = BORDER

# Итого Infrastructure
ws_infra.cell(row=25, column=1, value="ИТОГО ИНФРАСТРУКТУРА")
ws_infra.cell(row=25, column=4, value="=SUM(D5:D24)")
style_cell(ws_infra, 25, 1, TOTAL_FILL)
style_cell(ws_infra, 25, 4, TOTAL_FILL)
ws_infra['A25'].font = Font(bold=True)
ws_infra['D25'].font = Font(bold=True)
for col in range(1, 6):
    ws_infra.cell(row=25, column=col).border = BORDER

set_column_widths(ws_infra, [35, 18, 15, 15, 35])

# ==================== ЛИСТ 4: ТРАФИК ====================
ws_traffic = wb.create_sheet("Traffic")

ws_traffic['A1'] = "ТРАФИК И КОММУНИКАЦИИ"
ws_traffic.merge_cells('A1:E1')
ws_traffic['A1'].font = Font(bold=True, size=14, color="FFFFFF")
ws_traffic['A1'].fill = HEADER_FILL

headers = ["Канал", "Цена за событие ($)", "Количество", "Итого ($)", "Комментарий"]
for col, header in enumerate(headers, 1):
    ws_traffic.cell(row=3, column=col, value=header)
style_header(ws_traffic, 3, 5)

traffic_data = [
    ("УВЕДОМЛЕНИЯ", "", "", "", ""),
    ("Push Notifications", 0.0001, "=CONTROL!B22*10", "=B5*C5", "~10 пушей/пользователь"),
    ("Email (транзакционные)", 0.001, "=CONTROL!B22*3", "=B6*C6", "~3 email/пользователь"),
    ("SMS (критичные)", 0.05, "=CONTROL!B22*0.1", "=B7*C7", "10% пользователей"),
    ("", "", "", "", ""),
    ("API И ИНТЕГРАЦИИ", "", "", "", ""),
    ("Webhook calls", 0.0001, "=CONTROL!B22*5", "=B11*C11", "~5 webhooks/пользователь"),
    ("Payment Gateway (Stripe)", 0, 1, "=B12*C12", "% от транзакций отдельно"),
    ("Analytics (Mixpanel)", 0, 1, "=B13*C13", "Free tier"),
]

for i, (channel, price, qty, total, comment) in enumerate(traffic_data, 4):
    ws_traffic.cell(row=i, column=1, value=channel)
    ws_traffic.cell(row=i, column=2, value=price)
    ws_traffic.cell(row=i, column=3, value=qty)
    ws_traffic.cell(row=i, column=4, value=total)
    ws_traffic.cell(row=i, column=5, value=comment)

    if channel in ["УВЕДОМЛЕНИЯ", "API И ИНТЕГРАЦИИ"]:
        for col in range(1, 6):
            style_cell(ws_traffic, i, col, SECTION_FILL)

    for col in range(1, 6):
        ws_traffic.cell(row=i, column=col).border = BORDER

# Итого Traffic
ws_traffic.cell(row=15, column=1, value="ИТОГО ТРАФИК")
ws_traffic.cell(row=15, column=4, value="=SUM(D5:D14)")
style_cell(ws_traffic, 15, 1, TOTAL_FILL)
style_cell(ws_traffic, 15, 4, TOTAL_FILL)
ws_traffic['A15'].font = Font(bold=True)
ws_traffic['D15'].font = Font(bold=True)
for col in range(1, 6):
    ws_traffic.cell(row=15, column=col).border = BORDER

set_column_widths(ws_traffic, [35, 22, 15, 15, 35])

# ==================== ЛИСТ 5: ИТОГО ====================
ws_total = wb.create_sheet("TOTAL")

ws_total['A1'] = "СВОДКА РАСХОДОВ"
ws_total.merge_cells('A1:D1')
ws_total['A1'].font = Font(bold=True, size=16, color="FFFFFF")
ws_total['A1'].fill = HEADER_FILL

# Текущий месяц
ws_total['A3'] = "Расчет для месяца:"
ws_total['B3'] = "=CONTROL!B7"
ws_total['C3'] = "Пользователей:"
ws_total['D3'] = "=CONTROL!B22"
ws_total['B3'].fill = INPUT_FILL
ws_total['D3'].fill = TOTAL_FILL

headers = ["Категория", "Сумма ($)", "% от общего", "Комментарий"]
for col, header in enumerate(headers, 1):
    ws_total.cell(row=5, column=col, value=header)
style_header(ws_total, 5, 4)

total_data = [
    ("AI и Генерация", "=AI_Generation!D22", "=B6/$B$10*100", "Pixelcut, Video, LLM"),
    ("Инфраструктура", "=Infrastructure!D25", "=B7/$B$10*100", "Серверы, БД, Хранилище"),
    ("Трафик и Коммуникации", "=Traffic!D15", "=B8/$B$10*100", "Push, Email, SMS"),
]

for i, (cat, formula, pct, comment) in enumerate(total_data, 6):
    ws_total.cell(row=i, column=1, value=cat)
    ws_total.cell(row=i, column=2, value=formula)
    ws_total.cell(row=i, column=3, value=pct)
    ws_total.cell(row=i, column=4, value=comment)
    for col in range(1, 5):
        ws_total.cell(row=i, column=col).border = BORDER

# Общий итог
ws_total.cell(row=10, column=1, value="ОБЩИЙ ИТОГ")
ws_total.cell(row=10, column=2, value="=SUM(B6:B8)")
ws_total.cell(row=10, column=3, value="100%")
style_cell(ws_total, 10, 1, TOTAL_FILL)
style_cell(ws_total, 10, 2, TOTAL_FILL)
ws_total['A10'].font = Font(bold=True, size=12)
ws_total['B10'].font = Font(bold=True, size=12)
for col in range(1, 5):
    ws_total.cell(row=10, column=col).border = BORDER

# Unit Economics
ws_total['A12'] = "UNIT ECONOMICS"
ws_total.merge_cells('A12:D12')
style_cell(ws_total, 12, 1, HEADER_FILL)
ws_total['A12'].font = HEADER_FONT

unit_data = [
    ("Стоимость 1 пользователя", "=B10/CONTROL!B22", "$", "Средняя за месяц"),
    ("Стоимость 1 примерки", "=AI_Generation!D5/CONTROL!B23", "$", "Только Pixelcut"),
    ("Стоимость 1 видео", "=AI_Generation!D9/CONTROL!B24", "$", "Только генерация"),
    ("Себестоимость образа (полная)", "=(AI_Generation!D5+AI_Generation!D19+AI_Generation!D20)/CONTROL!B23", "$", "Try-on + обработка"),
]

for i, (metric, formula, unit, comment) in enumerate(unit_data, 13):
    ws_total.cell(row=i, column=1, value=metric)
    ws_total.cell(row=i, column=2, value=formula)
    ws_total.cell(row=i, column=3, value=unit)
    ws_total.cell(row=i, column=4, value=comment)
    style_cell(ws_total, i, 2, TOTAL_FILL)
    for col in range(1, 5):
        ws_total.cell(row=i, column=col).border = BORDER

set_column_widths(ws_total, [35, 20, 15, 35])

# ==================== ЛИСТ 6: ПРОГНОЗ 6 МЕСЯЦЕВ ====================
ws_forecast = wb.create_sheet("Forecast_6M")

ws_forecast['A1'] = "ПРОГНОЗ НА 6 МЕСЯЦЕВ"
ws_forecast.merge_cells('A1:H1')
ws_forecast['A1'].font = Font(bold=True, size=16, color="FFFFFF")
ws_forecast['A1'].fill = HEADER_FILL

headers = ["Показатель", "Месяц 1", "Месяц 2", "Месяц 3", "Месяц 4", "Месяц 5", "Месяц 6", "Рост"]
for col, header in enumerate(headers, 1):
    ws_forecast.cell(row=3, column=col, value=header)
style_header(ws_forecast, 3, 8)

# Формулы для прогноза
forecast_rows = [
    ("Пользователи",
     "=CONTROL!B5",
     "=B4*(1+CONTROL!$B$6/100)",
     "=C4*(1+CONTROL!$B$6/100)",
     "=D4*(1+CONTROL!$B$6/100)",
     "=E4*(1+CONTROL!$B$6/100)",
     "=F4*(1+CONTROL!$B$6/100)",
     "=G4/B4-1"),
    ("Примерок (всего)",
     "=B4*CONTROL!$B$10",
     "=C4*CONTROL!$B$10",
     "=D4*CONTROL!$B$10",
     "=E4*CONTROL!$B$10",
     "=F4*CONTROL!$B$10",
     "=G4*CONTROL!$B$10",
     "=G5/B5-1"),
    ("", "", "", "", "", "", "", ""),
    ("ЗАТРАТЫ ($)", "", "", "", "", "", "", ""),
    ("AI и Генерация",
     "=B5*0.1 + B4*CONTROL!$B$11*0.5 + B4*CONTROL!$B$12*0.005*2",
     "=C5*0.1 + C4*CONTROL!$B$11*0.5 + C4*CONTROL!$B$12*0.005*2",
     "=D5*0.1 + D4*CONTROL!$B$11*0.5 + D4*CONTROL!$B$12*0.005*2",
     "=E5*0.1 + E4*CONTROL!$B$11*0.5 + E4*CONTROL!$B$12*0.005*2",
     "=F5*0.1 + F4*CONTROL!$B$11*0.5 + F4*CONTROL!$B$12*0.005*2",
     "=G5*0.1 + G4*CONTROL!$B$11*0.5 + G4*CONTROL!$B$12*0.005*2",
     "=G8/B8-1"),
    ("Инфраструктура",
     "=45*CEILING(B4/CONTROL!$B$19,1)+60*CEILING(CEILING(B4/CONTROL!$B$19,1)/2,1)+40+15+26+20",
     "=45*CEILING(C4/CONTROL!$B$19,1)+60*CEILING(CEILING(C4/CONTROL!$B$19,1)/2,1)+40+15+26+20",
     "=45*CEILING(D4/CONTROL!$B$19,1)+60*CEILING(CEILING(D4/CONTROL!$B$19,1)/2,1)+40+15+26+20",
     "=45*CEILING(E4/CONTROL!$B$19,1)+60*CEILING(CEILING(E4/CONTROL!$B$19,1)/2,1)+40+15+26+20",
     "=45*CEILING(F4/CONTROL!$B$19,1)+60*CEILING(CEILING(F4/CONTROL!$B$19,1)/2,1)+40+15+26+20",
     "=45*CEILING(G4/CONTROL!$B$19,1)+60*CEILING(CEILING(G4/CONTROL!$B$19,1)/2,1)+40+15+26+20",
     "=G9/B9-1"),
    ("Трафик",
     "=B4*10*0.0001+B4*3*0.001+B4*0.1*0.05",
     "=C4*10*0.0001+C4*3*0.001+C4*0.1*0.05",
     "=D4*10*0.0001+D4*3*0.001+D4*0.1*0.05",
     "=E4*10*0.0001+E4*3*0.001+E4*0.1*0.05",
     "=F4*10*0.0001+F4*3*0.001+F4*0.1*0.05",
     "=G4*10*0.0001+G4*3*0.001+G4*0.1*0.05",
     "=G10/B10-1"),
    ("", "", "", "", "", "", "", ""),
    ("ИТОГО МЕСЯЦ",
     "=B8+B9+B10",
     "=C8+C9+C10",
     "=D8+D9+D10",
     "=E8+E9+E10",
     "=F8+F9+F10",
     "=G8+G9+G10",
     "=G12/B12-1"),
    ("Накопительно",
     "=B12",
     "=B13+C12",
     "=C13+D12",
     "=D13+E12",
     "=E13+F12",
     "=F13+G12",
     ""),
    ("", "", "", "", "", "", "", ""),
    ("UNIT ECONOMICS", "", "", "", "", "", "", ""),
    ("$/пользователь",
     "=B12/B4",
     "=C12/C4",
     "=D12/D4",
     "=E12/E4",
     "=F12/F4",
     "=G12/G4",
     "=G16/B16-1"),
    ("$/примерка",
     "=B8/B5",
     "=C8/C5",
     "=D8/D5",
     "=E8/E5",
     "=F8/F5",
     "=G8/G5",
     "=G17/B17-1"),
]

for i, row_data in enumerate(forecast_rows, 4):
    for j, value in enumerate(row_data, 1):
        cell = ws_forecast.cell(row=i, column=j, value=value)
        cell.border = BORDER

    # Стилизация
    if row_data[0] in ["ЗАТРАТЫ ($)", "UNIT ECONOMICS"]:
        for col in range(1, 9):
            style_cell(ws_forecast, i, col, SECTION_FILL)
    elif row_data[0] in ["ИТОГО МЕСЯЦ", "Накопительно"]:
        for col in range(1, 9):
            style_cell(ws_forecast, i, col, TOTAL_FILL)
        ws_forecast.cell(row=i, column=1).font = Font(bold=True)

# Форматирование процентов для колонки H
for row in [4, 5, 8, 9, 10, 12, 16, 17]:
    cell = ws_forecast.cell(row=row, column=8)
    cell.number_format = '0.0%'

set_column_widths(ws_forecast, [20, 14, 14, 14, 14, 14, 14, 12])

# ==================== ЛИСТ 7: КАЛЬКУЛЯТОР ТОКЕНОВ ====================
ws_tokens = wb.create_sheet("Token_Calculator")

ws_tokens['A1'] = "КАЛЬКУЛЯТОР СТОИМОСТИ ТОКЕНОВ"
ws_tokens.merge_cells('A1:F1')
ws_tokens['A1'].font = Font(bold=True, size=16, color="FFFFFF")
ws_tokens['A1'].fill = HEADER_FILL

headers = ["Провайдер / Модель", "Input ($/1M)", "Output ($/1M)", "Токенов/запрос", "Запросов", "Итого ($)"]
for col, header in enumerate(headers, 1):
    ws_tokens.cell(row=3, column=col, value=header)
style_header(ws_tokens, 3, 6)

token_data = [
    ("OPENAI", "", "", "", "", ""),
    ("GPT-4o", 2.50, 10.00, 2000, "=CONTROL!B25", "=(B5*C5/1000000+C5*D5/1000000)*E5"),
    ("GPT-4o-mini", 0.15, 0.60, 1500, "=CONTROL!B25*0.5", "=(B6*D6/1000000+C6*D6/1000000)*E6"),
    ("GPT-4-turbo", 10.00, 30.00, 2000, 0, "=(B7*D7/1000000+C7*D7/1000000)*E7"),
    ("", "", "", "", "", ""),
    ("ANTHROPIC", "", "", "", "", ""),
    ("Claude 3.5 Sonnet", 3.00, 15.00, 2000, 0, "=(B11*D11/1000000+C11*D11/1000000)*E11"),
    ("Claude 3 Haiku", 0.25, 1.25, 1500, 0, "=(B12*D12/1000000+C12*D12/1000000)*E12"),
    ("", "", "", "", "", ""),
    ("GOOGLE", "", "", "", "", ""),
    ("Gemini 1.5 Pro", 1.25, 5.00, 2000, 0, "=(B16*D16/1000000+C16*D16/1000000)*E16"),
    ("Gemini 1.5 Flash", 0.075, 0.30, 1500, 0, "=(B17*D17/1000000+C17*D17/1000000)*E17"),
    ("", "", "", "", "", ""),
    ("EMBEDDINGS", "", "", "", "", ""),
    ("text-embedding-3-small", 0.02, 0, 500, "=CONTROL!B25", "=B21*D21/1000000*E21"),
    ("text-embedding-3-large", 0.13, 0, 500, 0, "=B22*D22/1000000*E22"),
]

for i, row_data in enumerate(token_data, 4):
    for j, value in enumerate(row_data, 1):
        cell = ws_tokens.cell(row=i, column=j, value=value)
        cell.border = BORDER

    if row_data[0] in ["OPENAI", "ANTHROPIC", "GOOGLE", "EMBEDDINGS"]:
        for col in range(1, 7):
            style_cell(ws_tokens, i, col, SECTION_FILL)

# Итого токены
ws_tokens.cell(row=24, column=1, value="ИТОГО LLM ЗАТРАТЫ")
ws_tokens.cell(row=24, column=6, value="=SUM(F5:F23)")
style_cell(ws_tokens, 24, 1, TOTAL_FILL)
style_cell(ws_tokens, 24, 6, TOTAL_FILL)
ws_tokens['A24'].font = Font(bold=True)
ws_tokens['F24'].font = Font(bold=True)
for col in range(1, 7):
    ws_tokens.cell(row=24, column=col).border = BORDER

# Памятка
ws_tokens['A26'] = "ПАМЯТКА: 1M = 1,000,000 токенов. ~750 слов = ~1000 токенов"
ws_tokens['A26'].font = Font(italic=True, color="666666")

set_column_widths(ws_tokens, [25, 15, 15, 18, 15, 15])

# Сохраняем
wb.save('/workspace/project_cost_calculator.xlsx')
print("✅ Калькулятор создан: /workspace/project_cost_calculator.xlsx")
print("\nСтруктура файла:")
print("  1. CONTROL - Панель управления (вводные данные)")
print("  2. AI_Generation - Затраты на ИИ (Pixelcut, Video, LLM)")
print("  3. Infrastructure - Серверы и инфраструктура")
print("  4. Traffic - Трафик и коммуникации")
print("  5. TOTAL - Сводка и Unit Economics")
print("  6. Forecast_6M - Прогноз на 6 месяцев")
print("  7. Token_Calculator - Калькулятор токенов")
